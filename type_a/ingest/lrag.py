import os
import glob
import io
import asyncio
from pathlib import Path
import aiofiles

from lightrag import LightRAG
from lightrag.llm.ollama import ollama_model_complete, ollama_embed
from lightrag.utils import EmbeddingFunc
from lightrag.kg.shared_storage import initialize_pipeline_status
from lightrag.utils import logger

import pandas as pd
from docx import Document
from PyPDF2 import PdfReader
from markdown import markdown
import re
# ——— Extractor functions ———
async def extract_text(raw: bytes, path: Path) -> str:
    return raw.decode("utf-8", errors="ignore")

async def extract_md(raw: bytes, path: Path) -> str:
    # Simply return raw markdown text
    return raw.decode("utf-8", errors="ignore")

async def extract_pdf(raw: bytes, path: Path) -> str:
    from PyPDF2 import PdfReader
    buf = io.BytesIO(raw)
    reader = PdfReader(buf)
    pages = [p.extract_text() or "" for p in reader.pages]
    return "\n".join(pages)

async def extract_docx(raw: bytes, path: Path) -> str:
    from docx import Document
    buf = io.BytesIO(raw)
    doc = Document(buf)
    return "\n".join(p.text for p in doc.paragraphs)

async def extract_xlsx(raw: bytes, path: Path) -> str:
    from openpyxl import load_workbook
    buf = io.BytesIO(raw)
    wb = load_workbook(buf, read_only=True, data_only=True)
    lines = []
    for sheet in wb:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            lines.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(lines)

# ——— Registry mapping extensions to extractors ———
EXTRACTORS = {
    ".txt":  extract_text,
    ".md":   extract_md,
    ".pdf":  extract_pdf,
    ".docx": extract_docx,
    ".xlsx": extract_xlsx,
}

async def pipeline_enqueue_file(rag: LightRAG, file_path: Path) -> bool:
    """
    Read a file, extract its text, and enqueue for KG ingestion.
    Returns True on success, False otherwise.
    """
    ext = file_path.suffix.lower()
    extractor = EXTRACTORS.get(ext)
    if extractor is None:
        logger.warning(f"Skipping unsupported file type: {file_path.name}")
        return False

    try:
        # Read raw bytes once
        async with aiofiles.open(file_path, "rb") as f:
            raw = await f.read()

        # Extract content
        content = await extractor(raw, file_path)
        if not content.strip():
            logger.warning(f"No content extracted from {file_path.name}")
            return False

        #--Alternative--
        # Enqueue _and_ immediately process the full pipeline end-to-end
        # await rag.apipeline_process_enqueue_documents(
        #     [content], file_paths=[str(file_path)]
        #     )

        # Enqueue into LightRAG pipeline
        await rag.apipeline_enqueue_documents(
            [content], file_paths=[str(file_path)]
        )
        logger.info(f"Enqueued: {file_path.name}")
        return True

    except Exception as e:
        logger.error(f"Error processing {file_path.name}: {e}")
        return False

async def build_knowledge_graph(data_dir: str, storage_dir: str):
    """
    Initialize LightRAG, ingest all supported files in data_dir,
    and persist the KG to Neo4j under storage_dir.
    """
    # Ensure storage directory exists
    os.makedirs(storage_dir, exist_ok=True)

    # Initialize LightRAG with Neo4jStorage
    BATCH_SIZE_NODES = 500
    BATCH_SIZE_EDGES = 100
    os.environ.setdefault("NEO4J_URI", os.getenv("NEO4J_URI", "neo4j://localhost:7687"))
    os.environ.setdefault("NEO4J_USERNAME", os.getenv("NEO4J_USERNAME", "neo4j"))
    os.environ.setdefault("NEO4J_PASSWORD", os.getenv("NEO4J_PASSWORD", "mypassword"))
    
    # mongo
    os.environ.setdefault("MONGO_URI", os.getenv("MONGO_URI", "mongodb://mongoUser:mongoPass@localhost:27017/"))
    os.environ.setdefault("MONGO_DATABASE", os.getenv("MONGO_DATABASE", "LightRAG"))

    # pgvector
    os.environ.setdefault("POSTGRES_HOST", os.getenv("POSTGRES_HOST", "localhost"))
    os.environ.setdefault("POSTGRES_PORT", os.getenv("POSTGRES_PORT", "15432"))
    os.environ.setdefault("POSTGRES_USER", os.getenv("POSTGRES_USER", "demo"))
    os.environ.setdefault("POSTGRES_PASSWORD", os.getenv("POSTGRES_PASSWORD", "demo1234"))
    os.environ.setdefault("POSTGRES_DATABASE", os.getenv("POSTGRES_DATABASE", "rag"))

    rag = LightRAG(
        working_dir=storage_dir,
        llm_model_func=ollama_model_complete,
        llm_model_name="qwen2.5:latest",
        llm_model_max_async=4,
        llm_model_max_token_size=32768,
        llm_model_kwargs={
            "host": "http://127.0.0.1:11434",
            "options": {"num_ctx": 32768},
        },
        embedding_func=EmbeddingFunc(
            embedding_dim=4096,
            max_token_size=8192,
            func=lambda texts: ollama_embed(
                texts=texts, embed_model="bge-m3:latest", host="http://127.0.0.1:11434"
            ),
        ),
        kv_storage="MongoKVStorage",
        graph_storage="Neo4JStorage",
        vector_storage="PGVectorStorage",
        doc_status_storage="PGDocStatusStorage",
        # auto_manage_storages_states=False,
    )
    await rag.initialize_storages()
    await initialize_pipeline_status()

    # Discover files
    patterns = [f"*.{ext.lstrip('.')}" for ext in EXTRACTORS.keys()]
    for pattern in patterns:
        for filepath in glob.glob(os.path.join(data_dir, pattern)):
            print(filepath)
            await pipeline_enqueue_file(rag, Path(filepath))

    # Finalize and flush to Neo4j, postgres, mongo
    await rag.apipeline_process_enqueue_documents()

    await rag.finalize_storages()
    logger.info("Knowledge graph ingestion complete.")

if __name__ == "__main__":
    # import argparse

    # parser = argparse.ArgumentParser(
    #     description="Ingest documents into a LightRAG knowledge graph stored in Neo4j."
    # )
    # parser.add_argument(
    #     "--data-dir", required=True,
    #     help="Directory containing files to ingest (pdf, docx, xlsx, md, txt)."
    # )
    # parser.add_argument(
    #     "--storage-dir", required=True,
    #     help="Directory for LightRAG working/storage files."
    # )
    # args = parser.parse_args()
    # # print(args.data_dir, args.storage_dir)
    asyncio.run(build_knowledge_graph("./data", "./rag_storage"))
